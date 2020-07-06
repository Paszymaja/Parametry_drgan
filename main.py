import openpyxl
import numpy as np

wb = openpyxl.load_workbook('dane.xlsx')
arkusz = wb['Arkusz1']


def wczytanie_danych(sheet):
    tab = []
    for row_i in range(2, 62):
        for column_i in range(1, 8):
            x1 = float(sheet.cell(row_i, column_i).value)
            tab.append(x1)
    mat = np.array(tab).reshape(60, 7)
    return mat


def wczytanie_stanowisk(sheet):
    tab = []
    for row_i in range(2, 5):
        for column_i in range(9, 12):
            x1 = float(sheet.cell(row_i, column_i).value)
            tab.append(x1)
    mat = np.array(tab).reshape(3, 3)
    return mat


def macierz_a(macierz, macierz_stanowisk):
    tab = []

    def value(a, b):
        return np.sqrt(pow(macierz[i][3] - macierz_stanowisk[a][b], 2) +
                       pow(macierz[i][4] - macierz_stanowisk[a][b + 1], 2) +
                       pow(500, 2))

    for i in range(60):
        for j in range(4):
            if j == 0:
                tab.append(np.log10(macierz[i][2]))
            elif j == 1:
                if macierz[i][0] == 1:
                    tab.append(np.log10(value(0, 1)))
                elif macierz[i][0] == 2:
                    tab.append(np.log10(value(1, 1)))
                elif macierz[i][0] == 3:
                    tab.append(np.log10(value(2, 1)))
            elif j == 2:
                if macierz[i][0] == 1:
                    tab.append(value(0, 1))
                elif macierz[i][0] == 2:
                    tab.append(value(1, 1))
                elif macierz[i][0] == 3:
                    tab.append(value(2, 1))
            elif j == 3:
                tab.append(1)
    mat = np.array(tab).reshape(60, 4)
    return mat


def macierz_y(macierz):
    tab = [np.log10(macierz[i][5]) for i in range(60)]
    return np.array(tab).reshape(60, 1)


def odleglosci_epicentralne():
    data_list = np.arange(0, 2000, 500)
    data_list.tolist()
    return data_list


def macierz_y_prog(macierz, wyniki):
    tab = [(wyniki[0] * macierz[i][0])
           + (wyniki[1] * macierz[i][1])
           + (wyniki[2] * macierz[i][2])
           + wyniki[3]
           for i in range(60)]
    mat = np.array(tab).reshape(60, 1)
    return mat


def odchylenie_standardowe(macierz, macierz_prog):
    tab = [macierz[i][0] + macierz_prog[i][0] for i in range(60)]
    mat = np.array(tab).reshape(60, 1)
    return mat


def sigma(odchylenie, odchylenie_srednia):
    a = 0
    for i in range(60):
        a += pow(odchylenie[i] - odchylenie_srednia, 2)
    a = np.sqrt(a) / 59
    return a


def a_epicentralne(odleglosc, wyniki):
    tab = []
    for i in range(4):
        a = np.sqrt(pow(odleglosc[i], 2) + pow(500, 2))
        a = pow(10, (wyniki[0] * np.log10(200000)) + (wyniki[1] * np.log10(a)) + (wyniki[2] * a) + wyniki[3])
        tab.append(a)
    mat = np.array(tab).reshape(4, 1)
    return mat


def a_epicentralne_uf(a_epic, sigma_val):
    lam = 1.6707
    tab = [pow(10, np.log10(a_epic[i]) + (sigma_val * lam)) for i in range(4)]
    mat = np.array(tab).reshape(4, 1)
    return mat


def main():
    macierz_danych = wczytanie_danych(arkusz)
    stanowiska = wczytanie_stanowisk(arkusz)
    odl_ep = odleglosci_epicentralne()

    macierzA = macierz_a(macierz_danych, stanowiska)
    macierzY = macierz_y(macierz_danych)

    macierzT = macierzA.transpose()
    macierzA_mnozona = np.dot(macierzT, macierzA)
    macierzY_mnozona = np.dot(macierzT, macierzY)

    wynik = np.linalg.solve(macierzA_mnozona, macierzY_mnozona)
    macierzY_prog = macierz_y_prog(macierzA, wynik)
    odchylenie_stnd = odchylenie_standardowe(macierzY, macierzY_prog)
    odchylenie_stnd_srednia = sum(odchylenie_stnd) / 60
    sigma_wynik = sigma(odchylenie_stnd, odchylenie_stnd_srednia)
    A_epicentralne = a_epicentralne(odl_ep, wynik)
    A_epicentralne_UF = a_epicentralne_uf(A_epicentralne, sigma_wynik)

    print("Wynik")
    print(wynik)

    print('A_epicentralne')
    print(A_epicentralne)

    print('A_epicentralne_UF')
    print(A_epicentralne_UF)

    print('sigma_wynik')
    print(sigma_wynik)

    print('odchylenie_stnd_srednia')
    print(odchylenie_stnd_srednia)


if __name__ == '__main__':
    main()
